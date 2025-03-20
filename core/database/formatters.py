"""
Result formatters for database query results.

This module provides functions for formatting database query results
into various formats (JSON, CSV, XML, etc.).
"""

import logging
import csv
import json
import io
import xml.etree.ElementTree as ET
from typing import Dict, Any, List, Optional, Union, TextIO

logger = logging.getLogger(__name__)

def format_results(
    results: List[Dict[str, Any]], 
    format_type: str = 'json',
    options: Optional[Dict[str, Any]] = None
) -> Union[str, bytes, Dict[str, Any]]:
    """
    Format query results into the specified format.
    
    Args:
        results: Query results as a list of dictionaries
        format_type: Output format ('json', 'csv', 'xml', 'dict', etc.)
        options: Format-specific options
        
    Returns:
        Formatted results in the requested format
    """
    if not results:
        return "" if format_type != 'dict' else {}
    
    options = options or {}
    
    if format_type == 'json':
        return to_json(results, options)
    elif format_type == 'csv':
        return to_csv(results, options)
    elif format_type == 'xml':
        return to_xml(results, options)
    elif format_type == 'dict':
        return results  # Already in dictionary format
    else:
        logger.warning(f"Unsupported format type: {format_type}, returning as dict")
        return results


def to_json(
    results: List[Dict[str, Any]], 
    options: Dict[str, Any]
) -> str:
    """
    Convert results to JSON format.
    
    Args:
        results: Query results as list of dictionaries
        options: JSON formatting options
            - pretty: Whether to pretty-print the JSON (default: False)
            - ensure_ascii: Whether to escape non-ASCII characters (default: True)
            - root_name: Name of the root element for root-wrapped JSON (default: None)
    
    Returns:
        JSON string representation of the results
    """
    pretty = options.get('pretty', False)
    ensure_ascii = options.get('ensure_ascii', True)
    root_name = options.get('root_name')
    
    indent = 4 if pretty else None
    
    if root_name:
        data = {root_name: results}
    else:
        data = results
    
    try:
        return json.dumps(data, indent=indent, ensure_ascii=ensure_ascii, default=str)
    except Exception as e:
        logger.error(f"Error converting results to JSON: {str(e)}")
        return json.dumps({"error": "Failed to convert results to JSON"})


def to_csv(
    results: List[Dict[str, Any]], 
    options: Dict[str, Any]
) -> str:
    """
    Convert results to CSV format.
    
    Args:
        results: Query results as list of dictionaries
        options: CSV formatting options
            - delimiter: CSV delimiter (default: ',')
            - quotechar: CSV quote character (default: '"')
            - columns: List of columns to include (default: all)
            - include_header: Whether to include header row (default: True)
    
    Returns:
        CSV string representation of the results
    """
    if not results:
        return ""
    
    delimiter = options.get('delimiter', ',')
    quotechar = options.get('quotechar', '"')
    include_header = options.get('include_header', True)
    columns = options.get('columns')
    
    # If columns not specified, use keys from first result
    if not columns:
        columns = list(results[0].keys())
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, quotechar=quotechar, quoting=csv.QUOTE_MINIMAL)
    
    # Write header
    if include_header:
        writer.writerow(columns)
    
    # Write data rows
    for row in results:
        writer.writerow([row.get(col, '') for col in columns])
    
    return output.getvalue()


def to_xml(
    results: List[Dict[str, Any]], 
    options: Dict[str, Any]
) -> str:
    """
    Convert results to XML format.
    
    Args:
        results: Query results as list of dictionaries
        options: XML formatting options
            - root_name: Name of the root element (default: 'results')
            - row_name: Name of each row element (default: 'row')
            - pretty: Whether to pretty-print the XML (default: False)
    
    Returns:
        XML string representation of the results
    """
    root_name = options.get('root_name', 'results')
    row_name = options.get('row_name', 'row')
    pretty = options.get('pretty', False)
    
    # Create root element
    root = ET.Element(root_name)
    
    # Add rows
    for result in results:
        row_elem = ET.SubElement(root, row_name)
        
        # Add fields to row
        for key, value in result.items():
            # Skip None values
            if value is None:
                continue
                
            # Convert value to string
            if not isinstance(value, str):
                value = str(value)
            
            # Create element for this field
            # Replace spaces and special chars in key names
            safe_key = key.replace(' ', '_').replace(':', '_').replace('-', '_')
            field_elem = ET.SubElement(row_elem, safe_key)
            field_elem.text = value
    
    # Convert to string
    try:
        xml_str = ET.tostring(root, encoding='unicode')
        
        # Pretty print if requested
        if pretty:
            import xml.dom.minidom
            xml_str = xml.dom.minidom.parseString(xml_str).toprettyxml(indent="  ")
        
        return xml_str
    except Exception as e:
        logger.error(f"Error converting results to XML: {str(e)}")
        return f"<{root_name}><error>Failed to convert results to XML</error></{root_name}>"


def to_excel(
    results: List[Dict[str, Any]], 
    options: Dict[str, Any]
) -> bytes:
    """
    Convert results to Excel format (XLSX).
    
    Args:
        results: Query results as list of dictionaries
        options: Excel formatting options
            - sheet_name: Name of the worksheet (default: 'Data')
            - columns: List of columns to include (default: all)
            - include_header: Whether to include header row (default: True)
    
    Returns:
        Excel file content as bytes
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl module not found. Please install it to use Excel formatting.")
        return b""
    
    if not results:
        return b""
    
    sheet_name = options.get('sheet_name', 'Data')
    include_header = options.get('include_header', True)
    columns = options.get('columns')
    
    # If columns not specified, use keys from first result
    if not columns:
        columns = list(results[0].keys())
    
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write header
    if include_header:
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data rows
    start_row = 2 if include_header else 1
    for row_idx, row in enumerate(results, start_row):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ''))
    
    # Auto-size columns
    for col_idx, _ in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = min(adjusted_width, 50)  # Cap width at 50
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def write_to_file(
    results: List[Dict[str, Any]], 
    file_path: str,
    format_type: str = 'csv',
    options: Optional[Dict[str, Any]] = None
) -> bool:
    """
    Write query results to a file in the specified format.
    
    Args:
        results: Query results as list of dictionaries
        file_path: Path to output file
        format_type: Output format ('json', 'csv', 'xml', 'xlsx', etc.)
        options: Format-specific options
        
    Returns:
        True if file was written successfully, False otherwise
    """
    options = options or {}
    
    try:
        formatted_data = format_results(results, format_type, options)
        
        # Handle binary vs text output
        if format_type == 'xlsx':
            with open(file_path, 'wb') as f:
                f.write(formatted_data)
        else:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(formatted_data)
        
        logger.info(f"Successfully wrote results to {file_path}")
        return True
    
    except Exception as e:
        logger.error(f"Error writing results to file: {str(e)}")
        return False